import openpyxl

# file --> workbook --> sheets ---> rows ---> cells

file = "C:/Users/Mangesh/OneDrive/Desktop/PythonCollection.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row  # count number of rows in a Excel sheet
column = sheet.max_column  # count number of column in a Excel

# reading all the rows & column from Excel sheet
for r in range(1, rows+1):
    for c in range(1, column+1):
        print(sheet.cell(r, c).value, end=' ')
    print()
