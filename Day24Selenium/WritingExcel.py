import openpyxl

# same data

file = 'C:/Users/Mangesh/OneDrive/Desktop/TestData1.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # get active sheet from excel sheet = workbook["data"]

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "David"
sheet.cell(1, 3).value = "Designer"

sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "Smith"
sheet.cell(2, 3).value = "Developer"

sheet.cell(3, 1).value = 678
sheet.cell(3, 2).value = "john"
sheet.cell(3, 3).value = "Manager"
workbook.save(file)
